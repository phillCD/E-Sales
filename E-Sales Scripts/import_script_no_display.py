import base64
import os
import tempfile
from itertools import product

import openpyxl
import requests

box_url = 'http://localhost:8080/product-box'

display_url = 'http://localhost:8080/product-display'

product_url = 'http://localhost:8080/products'

product_to_api = {
    'name': '',
    'reference': '',
    'brand': '',
    'ncm': '',
    'cest': '',
    'ipi': '',
    'barcode': '',
    'height': '',
    'width': '',
    'weight': '',
    'length': '',
    'price': '',
}

box_to_api = {
    'product': {},
    'box_barcode': '',
    'box_weight': '',
    'box_height': '',
    'box_width': '',
    'box_length': '',
    'box_quantity': '',
}

display_to_api = {
    'product': {},
    'display_barcode': '',
    'display_weight': '',
    'display_height': '',
    'display_width': '',
    'display_length': '',
    'display_quantity': '',
}

def import_product(data):
    base64_data = data['file']
    decoded_data = base64.b64decode(base64_data)
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
        temp_file.write(decoded_data)
        temp_file_path = temp_file.name

    wb = openpyxl.load_workbook(temp_file_path)
    sheet = wb.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if any(cell is not None for cell in row):
            product_to_api['name'] = row[1]
            product_to_api['reference'] = row[0]
            product_to_api['brand'] = row[2]
            product_to_api['price'] = row[3]
            product_to_api['ipi'] = row[4]
            product_to_api['ncm'] = row[5]
            product_to_api['cest'] = row[6]
            product_to_api['barcode'] = row[7].replace('\xa0', '') if row[7] else ''
            product_to_api['length'] = row[8]
            product_to_api['width'] = row[9]
            product_to_api['height'] = row[10]
            product_to_api['weight'] = row[11]
            box_to_api['box_barcode'] = row[12].replace('\xa0', '') if row[12] else ''
            box_to_api['box_length'] = row[13]
            box_to_api['box_width'] = row[14]
            box_to_api['box_height'] = row[15]
            box_to_api['box_weight'] = row[16]
            box_to_api['box_quantity'] = row[17]
            print(product_to_api)
            print(box_to_api)

            response = requests.post(product_url, json=product_to_api)
            product_to_box = response.json()
            box_to_api['product'] = product_to_box
            response_box = requests.post(box_url, json=box_to_api)

            print(response_box.json())
            print(response.json())

    os.remove(temp_file_path)

def import_product_display(data):
    base64_data = data['file']
    decoded_data = base64.b64decode(base64_data)
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
        temp_file.write(decoded_data)
        temp_file_path = temp_file.name

    wb = openpyxl.load_workbook(temp_file_path)
    sheet = wb.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if any(cell is not None for cell in row):
            product_to_api['name'] = row[1]
            product_to_api['reference'] = row[0]
            product_to_api['price'] = None
            product_to_api['brand'] = row[3]
            product_to_api['ipi'] = row[4]
            product_to_api['ncm'] = row[5]
            product_to_api['cest'] = row[6]
            product_to_api['barcode'] = row[7].replace('\xa0', '') if row[7] else ''
            product_to_api['length'] = row[8]
            product_to_api['width'] = row[9]
            product_to_api['height'] = row[10]
            product_to_api['weight'] = row[11]
            box_to_api['box_barcode'] = row[18].replace('\xa0', '') if row[18] else ''
            box_to_api['box_length'] = row[19]
            box_to_api['box_width'] = row[20]
            box_to_api['box_height'] = row[21]
            box_to_api['box_weight'] = row[22]
            box_to_api['box_quantity'] = row[23]
            display_to_api['display_barcode'] = row[12].replace('\xa0', '') if row[12] else ''
            display_to_api['display_length'] = row[14]
            display_to_api['display_width'] = row[15]
            display_to_api['display_height'] = row[16]
            display_to_api['display_weight'] = row[17]
            display_to_api['display_quantity'] = row[13]
            print(product_to_api)
            print(box_to_api)
            print(display_to_api)

            response = requests.post(product_url, json=product_to_api)
            product_to_box = response.json()
            box_to_api['product'] = product_to_box
            response_box = requests.post(box_url, json=box_to_api)
            display_to_api['product'] = product_to_box
            response_display = requests.post(display_url, json=display_to_api)